from __future__ import annotations

import io
import math
import re
import zipfile
from datetime import date, datetime

from openpyxl import load_workbook

from .bundle import read_deterministic_bundle
from .models import DatasetSourceContract, ParsedDataset, QualityIssue
from .validation import RawPayloadError


SOURCE_ID = "transport.rail-station"
_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_MAX_XLSX_BYTES = 128 * 1024 * 1024
_MAX_UNCOMPRESSED_BYTES = 256 * 1024 * 1024
_MAX_COMPRESSION_RATIO = 100
_MAX_SHEETS = 10
_MAX_CELLS = 500_000
_REQUIRED_COLUMNS = {
    "철도운영기관명",
    "노선번호",
    "선명",
    "역번호",
    "역명",
    "도로명주소",
    "역위도",
    "역경도",
    "환승노선명",
    "데이터기준일자",
}


class RailStationAdapter:
    def parse(
        self,
        raw_bytes: bytes,
        contract: DatasetSourceContract,
        *,
        source_date: date | None,
    ) -> ParsedDataset:
        if contract.source_id != SOURCE_ID or source_date is None:
            raise RawPayloadError("rail source contract mismatch", "SOURCE_CONTRACT_MISMATCH")
        bundle = read_deterministic_bundle(
            raw_bytes,
            expected_source_id=SOURCE_ID,
            maximum_bytes=_MAX_XLSX_BYTES,
        )
        if bundle.temporal_value != source_date or len(bundle.artifacts) != 1:
            raise RawPayloadError("rail bundle metadata mismatch", "BUNDLE_MANIFEST_INVALID")
        artifact = bundle.artifacts[0]
        if artifact.media_type != _MEDIA_TYPE:
            raise RawPayloadError("rail artifact type mismatch", "BUNDLE_MANIFEST_INVALID")
        _inspect_xlsx_archive(artifact.content)
        try:
            workbook = load_workbook(
                io.BytesIO(artifact.content),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except Exception:
            raise RawPayloadError("rail XLSX cannot be parsed", "XLSX_INVALID") from None
        try:
            if not 1 <= len(workbook.worksheets) <= _MAX_SHEETS:
                raise RawPayloadError("rail XLSX sheet count is invalid", "XLSX_SIZE_LIMIT")
            rows: list[dict[str, object]] = []
            issues: list[QualityIssue] = []
            rejections: dict[int, tuple[str, ...]] = {}
            total_cells = 0
            for sheet in workbook.worksheets:
                iterator = sheet.iter_rows(values_only=True)
                header_values = next(iterator, None)
                if header_values is None:
                    continue
                headers = tuple(_clean(value) for value in header_values)
                if not _REQUIRED_COLUMNS.issubset(headers):
                    raise RawPayloadError("rail XLSX schema mismatch", "SOURCE_SCHEMA_MISMATCH")
                header_index = {header: index for index, header in enumerate(headers)}
                for values in iterator:
                    total_cells += len(values)
                    if total_cells > _MAX_CELLS:
                        raise RawPayloadError("rail XLSX cell limit exceeded", "XLSX_SIZE_LIMIT")
                    if not any(value is not None and str(value).strip() for value in values):
                        continue
                    provider_row = {
                        header: values[index] if index < len(values) else None
                        for header, index in header_index.items()
                    }
                    row_number = len(rows) + 1
                    normalized, reasons = _normalize(provider_row, source_date)
                    rows.append(normalized)
                    if reasons:
                        rejections[row_number] = tuple(reasons)
                        issues.extend(
                            QualityIssue(reason, "WARNING", row_number, {})
                            for reason in reasons
                        )
            if not rows:
                raise RawPayloadError("rail XLSX contains no data rows", "SOURCE_EMPTY")
            return ParsedDataset(
                rows=rows,
                issues=tuple(issues),
                row_rejections=rejections,
            )
        finally:
            workbook.close()


def _inspect_xlsx_archive(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            infos = archive.infolist()
            names = [info.filename for info in infos]
            if len(names) != len(set(names)):
                raise RawPayloadError("duplicate XLSX archive entry", "XLSX_INVALID")
            if any(name == "xl/vbaProject.bin" for name in names):
                raise RawPayloadError("XLSX macro is forbidden", "XLSX_MACRO_REJECTED")
            if any(name.startswith("xl/externalLinks/") for name in names):
                raise RawPayloadError(
                    "XLSX external links are forbidden", "XLSX_EXTERNAL_LINK_REJECTED"
                )
            total_uncompressed = sum(info.file_size for info in infos)
            if total_uncompressed > _MAX_UNCOMPRESSED_BYTES:
                raise RawPayloadError("XLSX expands beyond limit", "XLSX_SIZE_LIMIT")
            for info in infos:
                if info.file_size and info.compress_size == 0:
                    raise RawPayloadError("XLSX compression metadata is invalid", "XLSX_SIZE_LIMIT")
                if info.compress_size and info.file_size / info.compress_size > _MAX_COMPRESSION_RATIO:
                    raise RawPayloadError("XLSX compression ratio is unsafe", "XLSX_SIZE_LIMIT")
    except RawPayloadError:
        raise
    except zipfile.BadZipFile:
        raise RawPayloadError("XLSX archive is invalid", "XLSX_INVALID") from None


def _normalize(
    row: dict[str, object], source_date: date
) -> tuple[dict[str, object], list[str]]:
    operator = _clean(row.get("철도운영기관명"))
    line_number = _clean(row.get("노선번호"))
    line_name = _clean(row.get("선명"))
    station_number = _clean(row.get("역번호"))
    station_name = _clean(row.get("역명"))
    latitude = _number(row.get("역위도"))
    longitude = _number(row.get("역경도"))
    reasons: list[str] = []
    if not all((operator, line_number, station_number, station_name)):
        reasons.append("RAIL_STATION_IDENTITY_REQUIRED")
    if (
        latitude is None
        or longitude is None
        or not 32 <= latitude <= 39.5
        or not 124 <= longitude <= 132
    ):
        reasons.append("RAIL_STATION_COORDINATE_REQUIRED")
    row_date = _date(row.get("데이터기준일자"))
    if row_date != source_date:
        reasons.append("SOURCE_DATE_MIXED")
    transfer_lines = _transfer_lines(row.get("환승노선명"))
    return (
        {
            "station_occurrence_id": f"{operator}|{line_number}|{station_number}",
            "operator": operator,
            "line_number": line_number,
            "line_name": line_name,
            "station_number": station_number,
            "station_name": station_name,
            "road_address": _optional(row.get("도로명주소")),
            "latitude": latitude,
            "longitude": longitude,
            "transfer_lines": transfer_lines,
            "reference_date": source_date.isoformat(),
        },
        reasons,
    )


def _clean(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def _optional(value: object) -> str | None:
    value = _clean(value)
    return value or None


def _number(value: object) -> float | None:
    try:
        number = float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(_clean(value)[:10])
    except ValueError:
        return None


def _transfer_lines(value: object) -> list[str]:
    text = _clean(value)
    if not text:
        return []
    lines = [" ".join(part.split()) for part in re.split(r"[,;|]", text)]
    return list(dict.fromkeys(line for line in lines if line))
