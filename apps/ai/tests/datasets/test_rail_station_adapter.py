from __future__ import annotations

import io
import zipfile
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from ai_service.datasets import rail_station
from ai_service.datasets.bundle import (
    BundleArtifact,
    FileBundleArtifact,
    ReadArtifact,
    ReadBundle,
    build_deterministic_bundle,
    build_deterministic_bundle_file,
)
from ai_service.datasets.models import DatasetSourceContract
from ai_service.datasets.rail_station import RailStationAdapter
from ai_service.datasets.validation import RawPayloadError, validate_rows


HEADERS = [
    "철도운영기관명",
    "노선번호",
    "선명",
    "역번호",
    "역명",
    "도로명주소",
    "역위도",
    "역경도",
    "환승노선명",
    "데이터기준일자",
]

LIVE_HEADERS = [
    "역번호",
    "역사명",
    "노선번호",
    "노선명",
    "영문역사명",
    "한자역사명",
    "환승역구분",
    "환승노선번호",
    "환승노선명",
    "역위도",
    "역경도",
    "운영기관명",
    "역사도로명주소",
    "역사전화번호",
    "데이터기준일자",
]


def _contract() -> DatasetSourceContract:
    return DatasetSourceContract(
        source_id="transport.rail-station",
        provider="국토교통부",
        landing_url="https://www.data.go.kr/data/15013205/standard.do",
        acquisition_url="https://www.data.go.kr/data/15013205/standard.do",
        license_terms="공공누리 제1유형",
        attribution_requirements="출처 표시",
        license_reviewed_on=date(2026, 7, 19),
        refresh_frequency="연간",
        freshness_days=410,
        file_format="XLSX",
        encoding="binary",
        schema_version="rail-station-v1",
        coordinate_system="EPSG:4326",
        unique_key_fields=("station_occurrence_id",),
        required_fields=(
            "station_occurrence_id",
            "operator",
            "line_number",
            "line_name",
            "station_number",
            "station_name",
            "latitude",
            "longitude",
        ),
        expected_min_rows=1,
        expected_max_rows=10,
        maximum_row_change_ratio=0.1,
        maximum_rejected_ratio=0,
        contains_personal_data=False,
        owner="apps/ai",
    )


def _xlsx(rows: list[list[object]], *, headers: list[str] = HEADERS) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "도시철도역사"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _bundle(xlsx: bytes) -> bytes:
    return build_deterministic_bundle(
        source_id="transport.rail-station",
        endpoint_path="/data/15013205/standard.do",
        temporal_value=date(2026, 1, 1),
        artifacts=(
            BundleArtifact(
                "annual-release",
                "xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                xlsx,
            ),
        ),
    )


def test_same_station_name_on_multiple_lines_keeps_distinct_occurrences() -> None:
    rows = [
        ["서울교통공사", "02", "2호선", "201", "시청", "서울 중구", 37.5657, 126.977, "1호선", "2026-01-01"],
        ["서울교통공사", "01", "1호선", "132", "시청", "서울 중구", 37.5653, 126.977, "2호선", "2026-01-01"],
    ]

    parsed = RailStationAdapter().parse(
        _bundle(_xlsx(rows)), _contract(), source_date=date(2026, 1, 1)
    )

    assert len(parsed.rows) == 2
    assert len({row["station_occurrence_id"] for row in parsed.rows}) == 2
    assert parsed.rows[0]["transfer_lines"] == ["1호선"]
    assert parsed.rows[1]["transfer_lines"] == ["2호선"]
    assert parsed.row_rejections == {}


def test_same_line_number_with_different_line_names_keeps_distinct_occurrences() -> None:
    rows = [
        ["한국철도공사", "I4108", "경의중앙선", "1202", "상봉역", "서울 중랑구", 37.596754, 127.085553, "경춘선", "2025-04-08"],
        ["한국철도공사", "I4108", "경춘선", "1202", "상봉역", "서울 중랑구", 37.595254, 127.085853, "경의중앙선", "2025-12-31"],
    ]

    parsed = RailStationAdapter().parse(
        _bundle(_xlsx(rows)), _contract(), source_date=date(2026, 1, 1)
    )

    assert {row["station_occurrence_id"] for row in parsed.rows} == {
        "한국철도공사|I4108|경의중앙선|1202",
        "한국철도공사|I4108|경춘선|1202",
    }


def test_exact_occurrence_keeps_only_unique_latest_reference_row() -> None:
    rows = [
        ["서울교통공사", "S1107", "7호선", "0736", "총신대입구(이수)", "서울 동작구", 37.485258, 126.981766, "4호선", "2024-12-31"],
        ["서울교통공사", "S1107", "7호선", "0736", "이수", "서울 동작구", 37.485258, 126.981766, "4호선", "2025-12-31"],
    ]

    parsed = RailStationAdapter().parse(
        _bundle(_xlsx(rows)), _contract(), source_date=date(2026, 1, 1)
    )

    assert len(parsed.rows) == 1
    assert parsed.rows[0]["station_name"] == "이수"
    assert parsed.rows[0]["reference_date"] == "2025-12-31"
    assert [issue.reason_code for issue in parsed.issues] == [
        "RAIL_SUPERSEDED_OCCURRENCE"
    ]


def test_exact_occurrence_with_tied_latest_date_remains_blocking() -> None:
    rows = [
        ["서울교통공사", "S1107", "7호선", "0736", "이수", "서울 동작구", 37.485258, 126.981766, "4호선", "2025-12-31"],
        ["서울교통공사", "S1107", "7호선", "0736", "총신대입구(이수)", "서울 동작구", 37.485258, 126.981766, "4호선", "2025-12-31"],
    ]
    parsed = RailStationAdapter().parse(
        _bundle(_xlsx(rows)), _contract(), source_date=date(2026, 1, 1)
    )

    outcome = validate_rows(
        _contract(),
        parsed.rows,
        None,
        source_date=date(2026, 1, 1),
        collected_at=datetime(2026, 1, 2),
        adapter_issues=parsed.issues,
        adapter_rejections=parsed.row_rejections,
    )

    assert outcome.rejected_row_count == 1
    assert "DUPLICATE_UNIQUE_KEY" in {
        issue.reason_code for issue in outcome.issues
    }


def test_live_kric_headers_are_normalized_without_projecting_phone() -> None:
    row = [
        "201", "시청", "02", "2호선", "City Hall", "市廳", "환승역", "01",
        "1호선", 37.5657, 126.977, "서울교통공사", "서울 중구", "02-0000-0000",
        "2026-06-30",
    ]

    parsed = RailStationAdapter().parse(
        _bundle(_xlsx([row], headers=LIVE_HEADERS)),
        _contract(),
        source_date=date(2026, 1, 1),
    )

    assert parsed.rows[0]["station_occurrence_id"] == "서울교통공사|02|2호선|201"
    assert parsed.rows[0]["station_name"] == "시청"
    assert parsed.rows[0]["line_name"] == "2호선"
    assert parsed.rows[0]["road_address"] == "서울 중구"
    assert parsed.rows[0]["reference_date"] == "2026-06-30"
    assert parsed.row_rejections == {}
    assert "phone" not in parsed.rows[0]


def test_rail_file_adapter_streams_bundle_artifact(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    xlsx_path = tmp_path / "release.xlsx"
    xlsx_path.write_bytes(
        _xlsx(
            [[
                "서울교통공사", "02", "2호선", "201", "시청", "서울 중구",
                37.5657, 126.977, "1호선", "2026-01-01",
            ]]
        )
    )
    xlsx_path.chmod(0o600)
    bundle_path = tmp_path / "bundle.zip"
    bundle_path.touch(mode=0o600)
    build_deterministic_bundle_file(
        source_id="transport.rail-station",
        endpoint_path="/data/15013205/standard.do",
        temporal_value=date(2026, 1, 1),
        artifacts=(
            FileBundleArtifact(
                "annual-release",
                "xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                xlsx_path,
            ),
        ),
        target=bundle_path,
    )
    original_read = zipfile.ZipFile.read

    def reject_outer_artifact_read(archive, name, *args, **kwargs):
        if Path(archive.filename) == bundle_path and str(name).startswith("artifacts/"):
            raise AssertionError("prepared artifact must be streamed")
        return original_read(archive, name, *args, **kwargs)

    monkeypatch.setattr(Path, "read_bytes", lambda _path: (_ for _ in ()).throw(
        AssertionError("prepared bundle must not be materialized")
    ))
    monkeypatch.setattr(zipfile.ZipFile, "read", reject_outer_artifact_read)

    parsed = RailStationAdapter().parse_file(
        bundle_path, _contract(), source_date=date(2026, 1, 1)
    )

    assert parsed.rows[0]["station_occurrence_id"] == "서울교통공사|02|2호선|201"


def test_rail_file_adapter_fails_closed_on_contract_metadata_and_parser(
    tmp_path: Path,
) -> None:
    adapter = RailStationAdapter()
    with pytest.raises(RawPayloadError) as error:
        adapter.parse_file(tmp_path / "missing.zip", _contract(), source_date=None)
    assert error.value.reason_code == "SOURCE_CONTRACT_MISMATCH"

    xlsx_path = tmp_path / "release.xlsx"
    xlsx_path.write_bytes(_xlsx([]))
    xlsx_path.chmod(0o600)
    wrong_media_bundle = tmp_path / "wrong-media.zip"
    wrong_media_bundle.touch(mode=0o600)
    build_deterministic_bundle_file(
        source_id="transport.rail-station",
        endpoint_path="/data/15013205/standard.do",
        temporal_value=date(2026, 1, 1),
        artifacts=(
            FileBundleArtifact("annual-release", "xlsx", "text/plain", xlsx_path),
        ),
        target=wrong_media_bundle,
    )
    with pytest.raises(RawPayloadError) as error:
        adapter.parse_file(
            wrong_media_bundle, _contract(), source_date=date(2026, 1, 1)
        )
    assert error.value.reason_code == "BUNDLE_MANIFEST_INVALID"

    malformed_xlsx = tmp_path / "malformed.xlsx"
    with zipfile.ZipFile(malformed_xlsx, "w") as archive:
        archive.writestr("placeholder", b"not-an-openxml-workbook")
    malformed_xlsx.chmod(0o600)
    malformed_bundle = tmp_path / "malformed-bundle.zip"
    malformed_bundle.touch(mode=0o600)
    build_deterministic_bundle_file(
        source_id="transport.rail-station",
        endpoint_path="/data/15013205/standard.do",
        temporal_value=date(2026, 1, 1),
        artifacts=(
            FileBundleArtifact(
                "annual-release",
                "xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                malformed_xlsx,
            ),
        ),
        target=malformed_bundle,
    )
    with pytest.raises(RawPayloadError) as error:
        adapter.parse_file(
            malformed_bundle, _contract(), source_date=date(2026, 1, 1)
        )
    assert error.value.reason_code == "XLSX_INVALID"


def test_missing_station_location_is_blocking_row_rejection() -> None:
    rows = [
        ["서울교통공사", "02", "2호선", "201", "시청", "서울 중구", None, None, "1호선", "2026-01-01"]
    ]

    parsed = RailStationAdapter().parse(
        _bundle(_xlsx(rows)), _contract(), source_date=date(2026, 1, 1)
    )

    assert parsed.row_rejections == {1: ("RAIL_STATION_COORDINATE_REQUIRED",)}


@pytest.mark.parametrize(
    ("entry_name", "reason_code"),
    [
        ("xl/vbaProject.bin", "XLSX_MACRO_REJECTED"),
        ("xl/externalLinks/externalLink1.xml", "XLSX_EXTERNAL_LINK_REJECTED"),
    ],
)
def test_unsafe_xlsx_parts_are_rejected(entry_name: str, reason_code: str) -> None:
    source = _xlsx(
        [["서울교통공사", "02", "2호선", "201", "시청", "서울 중구", 37.5657, 126.977, "", "2026-01-01"]]
    )
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(source)) as original, zipfile.ZipFile(output, "w") as changed:
        for entry in original.infolist():
            changed.writestr(entry, original.read(entry.filename))
        changed.writestr(entry_name, b"unsafe")

    with pytest.raises(RawPayloadError) as error:
        RailStationAdapter().parse(
            _bundle(output.getvalue()), _contract(), source_date=date(2026, 1, 1)
        )

    assert error.value.reason_code == reason_code


def test_source_contract_and_bundle_temporal_values_fail_closed() -> None:
    with pytest.raises(RawPayloadError) as error:
        RailStationAdapter().parse(_bundle(_xlsx([])), _contract(), source_date=None)
    assert error.value.reason_code == "SOURCE_CONTRACT_MISMATCH"

    with pytest.raises(RawPayloadError) as error:
        RailStationAdapter().parse(
            _bundle(_xlsx([])), _contract(), source_date=date(2026, 1, 2)
        )
    assert error.value.reason_code == "BUNDLE_MANIFEST_INVALID"


def test_schema_mismatch_fails_closed() -> None:
    workbook = Workbook()
    workbook.active.append(["역명"])
    workbook.active.append(["시청"])
    output = io.BytesIO()
    workbook.save(output)

    with pytest.raises(RawPayloadError) as error:
        RailStationAdapter().parse(
            _bundle(output.getvalue()), _contract(), source_date=date(2026, 1, 1)
        )

    assert error.value.reason_code == "SOURCE_SCHEMA_MISMATCH"


def test_identity_and_coordinate_rejections_preserve_row_provenance_date() -> None:
    rows = [
        ["", "", "2호선", "", "시청", "서울 중구", 91, 181, "2호선, 2호선;1호선", "2026-01-02"]
    ]

    parsed = RailStationAdapter().parse(
        _bundle(_xlsx(rows)), _contract(), source_date=date(2026, 1, 1)
    )

    assert parsed.row_rejections[1] == (
        "RAIL_STATION_IDENTITY_REQUIRED",
        "RAIL_STATION_COORDINATE_REQUIRED",
    )
    assert parsed.rows[0]["transfer_lines"] == ["2호선", "1호선"]
    assert parsed.rows[0]["reference_date"] == "2026-01-02"


def test_blank_or_invalid_row_date_is_nullable_provenance() -> None:
    rows = [
        ["서울교통공사", "02", "2호선", "201", "시청", "서울 중구", 37.5657, 126.977, "", ""],
        ["서울교통공사", "02", "2호선", "202", "을지로입구", "서울 중구", 37.566, 126.982, "", "1900-01-00"],
    ]

    parsed = RailStationAdapter().parse(
        _bundle(_xlsx(rows)), _contract(), source_date=date(2026, 1, 1)
    )

    assert [row["reference_date"] for row in parsed.rows] == [None, None]
    assert parsed.row_rejections == {}
    assert [(issue.reason_code, issue.severity, issue.row_number) for issue in parsed.issues] == [
        ("RAIL_ROW_REFERENCE_DATE_INVALID", "WARNING", 2),
    ]


def test_empty_workbook_fails_closed() -> None:
    with pytest.raises(RawPayloadError) as error:
        RailStationAdapter().parse(
            _bundle(_xlsx([])), _contract(), source_date=date(2026, 1, 1)
        )

    assert error.value.reason_code == "SOURCE_EMPTY"


def test_xlsx_media_type_and_archive_fail_closed(monkeypatch) -> None:
    bundle = ReadBundle(
        source_id="transport.rail-station",
        complete=True,
        endpoint_path="/data/15013205/standard.do",
        temporal_value=date(2026, 1, 1),
        artifacts=(ReadArtifact("release", "text/plain", b"bad"),),
    )
    monkeypatch.setattr(rail_station, "read_deterministic_bundle", lambda *_args, **_kwargs: bundle)
    with pytest.raises(RawPayloadError) as error:
        RailStationAdapter().parse(b"raw", _contract(), source_date=date(2026, 1, 1))
    assert error.value.reason_code == "BUNDLE_MANIFEST_INVALID"

    bundle = ReadBundle(
        source_id="transport.rail-station",
        complete=True,
        endpoint_path="/data/15013205/standard.do",
        temporal_value=date(2026, 1, 1),
        artifacts=(ReadArtifact("release", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", b"bad"),),
    )
    monkeypatch.setattr(rail_station, "read_deterministic_bundle", lambda *_args, **_kwargs: bundle)
    with pytest.raises(RawPayloadError) as error:
        RailStationAdapter().parse(b"raw", _contract(), source_date=date(2026, 1, 1))
    assert error.value.reason_code == "XLSX_INVALID"


def test_duplicate_xlsx_archive_entry_is_rejected(monkeypatch) -> None:
    output = io.BytesIO()
    with pytest.warns(UserWarning, match="Duplicate name"):
        with zipfile.ZipFile(output, "w") as archive:
            archive.writestr("same", b"one")
            archive.writestr("same", b"two")
    bundle = ReadBundle(
        source_id="transport.rail-station",
        complete=True,
        endpoint_path="/data/15013205/standard.do",
        temporal_value=date(2026, 1, 1),
        artifacts=(ReadArtifact("release", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", output.getvalue()),),
    )
    monkeypatch.setattr(rail_station, "read_deterministic_bundle", lambda *_args, **_kwargs: bundle)

    with pytest.raises(RawPayloadError) as error:
        RailStationAdapter().parse(b"raw", _contract(), source_date=date(2026, 1, 1))

    assert error.value.reason_code == "XLSX_INVALID"


def test_sheet_count_limit_is_enforced() -> None:
    workbook = Workbook()
    for index in range(10):
        workbook.create_sheet(f"extra-{index}")
    output = io.BytesIO()
    workbook.save(output)

    with pytest.raises(RawPayloadError) as error:
        RailStationAdapter().parse(
            _bundle(output.getvalue()), _contract(), source_date=date(2026, 1, 1)
        )

    assert error.value.reason_code == "XLSX_SIZE_LIMIT"


def test_native_date_cells_and_empty_transfer_are_supported() -> None:
    rows = [
        ["서울교통공사", "02", "2호선", "201", "시청", "서울 중구", 37.5657, 126.977, "", datetime(2026, 1, 1)],
    ]

    parsed = RailStationAdapter().parse(
        _bundle(_xlsx(rows)), _contract(), source_date=date(2026, 1, 1)
    )

    assert parsed.rows[0]["transfer_lines"] == []
    assert parsed.row_rejections == {}
